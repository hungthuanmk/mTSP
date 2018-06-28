from routemanager import *
from openpyxl import load_workbook
from dustbin import *

def load_data(file_name, first_row, last_row):
    wb = load_workbook(file_name)
    # print(wb.get_sheet_names())
    ws = wb.get_sheet_by_name('address')
    # print(ws)

    id = 1
    for row in range(first_row, last_row+1):
        code = ws.cell(row=row, column=1).value
        address = ws.cell(row=row, column=3).value
        # print(code, ' | ', address)

        # node = Node(code=code, address=address)

        RouteManager.addDustbin(Dustbin(_id = id, _address=address))
        id += 1

        # nodes.append(node)

        # visualizer = Visualizer()
        # visualizer.draw_node(node)

    # visualizer.show()


def save_data(file_name, sheet_name, start_row, start_column, truck1, truck2, truck3):
    wb = load_workbook('Database.xlsx')
    if wb.get_sheet_names().count(sheet_name) == 0:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb.get_sheet_by_name(sheet_name)

    for row in range(len(truck1)):
        ws.cell(row=start_row + row, column=start_column).value = truck1[row]

    for row in range(len(truck2)):
        ws.cell(row=start_row + row, column=start_column+1).value = truck2[row]

    for row in range(len(truck3)):
        ws.cell(row=start_row + row, column=start_column+2).value = truck3[row]

    wb.save(file_name)



def export_distance(file_name, distances):
    wb = load_workbook(file_name)
    if wb.get_sheet_names().count('distance') == 0:
        ws = wb.create_sheet('distance')
    else:
        ws = wb.get_sheet_by_name('distance')

    keys = distances.keys()

    for str in keys:
        a, b = str.split('|')
        a = int(a)
        b = int(b)
        ws.cell(row=a, column=a).value = 0
        ws.cell(row=a, column=b).value = distances[str]

    wb.save(file_name)
